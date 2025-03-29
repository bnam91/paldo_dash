import os
from googleapiclient.discovery import build
import pandas as pd
from auth import get_credentials
from dotenv import load_dotenv
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
import tkinter as tk
from tkinter import ttk, filedialog
from datetime import datetime
import sys
from openpyxl.utils import get_column_letter
import shutil  # 파일 복사를 위한 모듈 추가
import openpyxl  # 엑셀 파일 편집을 위한 모듈
import subprocess  # 폴더 열기 위한 모듈 추가

class StyleConverterGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("스타일 변환기")
        self.root.geometry("400x700")
        
        # 저장 경로 표시 및 선택
        tk.Label(root, text="저장 경로:").pack(pady=5)
        self.path_frame = tk.Frame(root)
        self.path_frame.pack(fill='x', padx=5)
        
        self.path_entry = tk.Entry(self.path_frame)
        self.path_entry.pack(side='left', fill='x', expand=True)
        
        # 폴더 열기 버튼을 먼저 배치
        tk.Button(self.path_frame, text="폴더 열기", command=self.open_folder).pack(side='right', padx=5)
        
        # 경로 선택 버튼을 그 다음에 배치
        tk.Button(self.path_frame, text="경로 선택", command=self.select_path).pack(side='right', padx=5)
        
        # 시트명 입력
        tk.Label(root, text="(정확히 입력)시트명:").pack(pady=5)
        self.sheet_name = tk.Entry(root)
        self.sheet_name.pack(pady=5)
        
        # 제품명 입력
        tk.Label(root, text="(정확히 입력)제품명:").pack(pady=5)
        self.product_name = tk.Entry(root)
        self.product_name.pack(pady=5)
        
        # 플랫폼 선택 (콤보박스로 변경)
        tk.Label(root, text="플랫폼 선택:").pack(pady=5)
        
        # 플랫폼 콤보박스와 입력창을 위한 프레임
        platform_frame = tk.Frame(root)
        platform_frame.pack(fill='x', padx=20)
        
        # 플랫폼 콤보박스 옵션
        self.platform_options = ["쿠팡", "알리", "마켓컬리", "카카오", "직접입력"]
        self.platform_var = tk.StringVar(value=self.platform_options[0])  # 기본값 설정
        
        self.platform_combo = ttk.Combobox(platform_frame, textvariable=self.platform_var, 
                                           values=self.platform_options, width=15)
        self.platform_combo.pack(side='left', fill='x', expand=True)
        
        # 직접 입력용 입력창
        self.platform_entry = tk.Entry(platform_frame, width=15)
        self.platform_entry.pack(side='left', padx=(5, 0), fill='x', expand=True)
        self.platform_entry.pack_forget()  # 처음에는 숨김
        
        # 콤보박스 선택 변경 시 이벤트 처리
        self.platform_combo.bind("<<ComboboxSelected>>", self.toggle_platform_input)
        
        # 차감내역 캡쳐 체크박스 추가 (간격 추가)
        checkbox_frame = tk.Frame(root)
        checkbox_frame.pack(pady=15)  # 위아래 간격 추가
        
        self.capture_var = tk.BooleanVar(value=True)  # 기본값 True로 설정
        self.capture_checkbox = tk.Checkbutton(checkbox_frame, 
                                             text="차감내역 생성",
                                             variable=self.capture_var)
        self.capture_checkbox.pack()
        
        # 변환 버튼
        tk.Button(root, text="스타일 변환하기", command=self.convert_style).pack(pady=20)
        
        # 보고메일 관련 프레임 생성
        report_frame = tk.Frame(root)
        report_frame.pack(pady=10, fill='x', padx=20)  # 좌우 여백 추가
        
        # 담당자 라벨 (별도 프레임에 배치)
        tk.Label(root, text="담당자:").pack(pady=(10, 5))  # 위쪽 여백 추가
        
        # 담당자 콤보박스와 입력창을 위한 프레임
        recipient_frame = tk.Frame(root)
        recipient_frame.pack(fill='x', padx=20)  # 좌우 여백 추가
        
        # 담당자 콤보박스 옵션
        self.recipient_options = ["배세웅 책임", "김동락 책임", "유호경 선임", "조민우 선임", "직접입력"]
        self.recipient_var = tk.StringVar(value=self.recipient_options[0])  # 기본값 설정
        
        self.recipient_combo = ttk.Combobox(recipient_frame, textvariable=self.recipient_var, 
                                           values=self.recipient_options, width=15)
        self.recipient_combo.pack(side='left', fill='x', expand=True)
        
        # 직접 입력용 입력창
        self.recipient_entry = tk.Entry(recipient_frame, width=15)
        self.recipient_entry.pack(side='left', padx=(5, 0), fill='x', expand=True)
        self.recipient_entry.pack_forget()  # 처음에는 숨김
        
        # 콤보박스 선택 변경 시 이벤트 처리
        self.recipient_combo.bind("<<ComboboxSelected>>", self.toggle_recipient_input)
        
        # 보고메일 버튼 추가 (별도 프레임에 배치)
        button_frame = tk.Frame(root)
        button_frame.pack(pady=10, fill='x', padx=20)  # 좌우 여백 추가
        
        self.report_email_button = tk.Button(button_frame, text="보고메일 양식 보기", 
                                            command=self.show_report_email_from_button)
        self.report_email_button.pack(fill='x')  # 버튼을 프레임 너비에 맞춤
        
        # 담당자 입력 필드 변경 시 보고메일 버튼 상태 업데이트
        self.recipient_entry.bind('<KeyRelease>', self.update_report_button_state)
        
        # 초기 버튼 상태 설정
        self.update_report_button_state()
        
        # 상태 메시지
        self.status_label = tk.Label(root, text="")
        self.status_label.pack(pady=10)

    def toggle_platform_input(self, event=None):
        """플랫폼 콤보박스에서 '직접입력' 선택 시 입력창 표시/숨김"""
        if self.platform_var.get() == "직접입력":
            self.platform_entry.pack(side='left', padx=(5, 0), fill='x', expand=True)
            self.platform_entry.focus()  # 입력창에 포커스 설정
        else:
            self.platform_entry.pack_forget()

    def toggle_recipient_input(self, event=None):
        """담당자 콤보박스에서 '직접입력' 선택 시 입력창 표시/숨김"""
        if self.recipient_var.get() == "직접입력":
            self.recipient_entry.pack(side='left', padx=(5, 0), fill='x', expand=True)
            self.recipient_entry.focus()  # 입력창에 포커스 설정
        else:
            self.recipient_entry.pack_forget()

    def select_path(self):
        directory = filedialog.askdirectory(
            initialdir=self.path_entry.get(),
            title="저장할 폴더를 선택하세요"
        )
        if directory:
            self.path_entry.delete(0, tk.END)
            self.path_entry.insert(0, directory)

    def open_folder(self):
        """저장 경로 폴더를 여는 함수"""
        path = self.path_entry.get()
        if os.path.exists(path):
            # 운영체제에 따라 폴더 열기 명령 실행
            if sys.platform == 'win32':  # Windows
                os.startfile(path)
            elif sys.platform == 'darwin':  # macOS
                subprocess.run(['open', path])
            else:  # Linux
                subprocess.run(['xdg-open', path])
        else:
            self.status_label.config(text="저장 경로가 존재하지 않습니다.")

    def convert_style(self):
        sheet_name = self.sheet_name.get()
        product = self.product_name.get()
        platform = self.platform_entry.get() if self.platform_var.get() == "직접입력" else self.platform_var.get()
        save_path = self.path_entry.get()
        
        # 경로가 비어있는지 먼저 확인
        if not save_path:
            self.status_label.config(text="저장 경로를 지정해주세요.")
            return
        
        # 경로가 존재하는지 확인
        if not os.path.exists(save_path):
            self.status_label.config(text="지정한 저장 경로가 존재하지 않습니다.")
            return
        
        if not all([sheet_name, product, platform]):
            self.status_label.config(text="모든 필드를 입력해주세요.")
            return
        
        # 파일명 생성
        today = datetime.now().strftime("%y%m%d")
        filename = f"현황판_{product}_{platform}리뷰체험단_고야앤드_{today}"
        
        try:
            # product와 platform 정보를 함께 전달
            self.download_sheet_to_excel(sheet_name, filename, product, platform, save_path)
            
            # 차감내역 체크박스가 체크되어 있으면 견적서 샘플 복사
            if self.capture_var.get():
                self.copy_invoice_template(product, save_path)
                
            self.status_label.config(text="변환 완료!")
            
            # 저장 경로 열기
            self.open_folder()
            
            # 변환 완료 후 보고메일 양식 팝업 표시
            self.show_report_email(product, platform, self.get_num_people())
        except Exception as e:
            self.status_label.config(text=f"오류 발생: {str(e)}")

    def download_sheet_to_excel(self, sheet_name, filename, product_name, platform, save_path):
        # 스프레드시트 ID 추출 (URL에서 추출)
        SPREADSHEET_ID = '1CK2UXTy7HKjBe2T0ovm5hfzAAKZxZAR_ev3cbTPOMPs'
        
        # Google Sheets API 서비스 생성
        creds = get_credentials()
        service = build('sheets', 'v4', credentials=creds)
        
        try:
            # 시트 데이터 가져오기
            result = service.spreadsheets().values().get(
                spreadsheetId=SPREADSHEET_ID,
                range=sheet_name
            ).execute()
            
            values = result.get('values', [])
            
            if not values:
                print('데이터를 찾을 수 없습니다.')
                return
            
            # 5-7행을 빈 데이터로 채우기 (3행만 빈 데이터로)
            empty_rows = [[''] * len(values[7]) for _ in range(3)]
            
            # 실제 데이터의 최대 컬럼 수 확인
            max_cols = max(len(row) for row in values[8:])
            
            # 8번째 행이 실제 데이터보다 컬럼이 적은 경우 추가 컬럼 생성
            headers = values[7]
            if len(headers) < max_cols:
                for i in range(len(headers), max_cols):
                    headers.append(f'추가컬럼_{i+1}')
            
            # DataFrame으로 변환 (빈 행 + 헤더 + 데이터)
            all_data = empty_rows + [headers] + values[8:]
            df = pd.DataFrame(all_data)
            
            # B8 셀에 '번호' 입력 (빈 행이 3개이므로 인덱스 3의 1번 열에 해당)
            df.iloc[3, 1] = '번호'
            
            # 특정 열 삭제 (A, D, G, H, I, J열과 Q열 이후)
            columns_to_drop = [0, 3, 6, 7, 8, 9]
            if len(df.columns) > 16:
                columns_to_drop.extend(range(16, len(df.columns)))
            df = df.drop(df.columns[columns_to_drop], axis=1)
            
            # 헤더 순서 확인
            expected_headers = ['번호', '날짜', '제품 및 내역', '이름', '금액', '구매여부', '포토리뷰', '리뷰작성', '캡쳐여부', '비고']
            actual_headers = df.iloc[3].tolist()  # 4번째 행(인덱스 3)이 헤더
            
            # 헤더 검증
            if len(actual_headers) != len(expected_headers):
                print('경고: 예상된 헤더 수와 실제 헤더 수가 다릅니다.')
                print(f'예상된 헤더: {expected_headers}')
                print(f'실제 헤더: {actual_headers}')
            else:
                for expected, actual in zip(expected_headers, actual_headers):
                    if expected != actual:
                        print(f'경고: 헤더 불일치 - 예상: {expected}, 실제: {actual}')
            
            # 저장 경로 설정
            output_file = os.path.join(save_path, f"{filename}.xlsx")
            
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name=sheet_name, header=False)
                
                worksheet = writer.sheets[sheet_name]
                
                # 모든 열의 너비를 13으로 설정
                for col in worksheet.columns:
                    col_letter = col[0].column_letter
                    worksheet.column_dimensions[col_letter].width = 13
                
                # 스타일 정의
                navy_color = '{:02x}{:02x}{:02x}'.format(66, 133, 244).upper()
                header_fill = PatternFill(start_color=navy_color, end_color=navy_color, fill_type='solid')
                white_bold_font = Font(color='FFFFFF', bold=True, name='맑은 고딕', size=11)
                
                # '제품 및 내역' 열 찾기
                product_col = None
                num_people = None
                
                for col in range(1, worksheet.max_column + 1):
                    if worksheet.cell(row=4, column=col).value == '제품 및 내역':
                        product_col = col
                        # 인원수 계산 (데이터가 있는 행의 수 계산)
                        num_people = 0
                        for r in range(5, worksheet.max_row + 1):
                            if worksheet.cell(row=r, column=product_col).value:
                                num_people += 1
                        break
                
                if not num_people:
                    num_people = 0  # 기본값 설정
                
                # 인원수를 클래스 변수로 저장
                self.num_people = num_people
                
                # A2-J2 병합 및 스타일 적용
                worksheet.merge_cells('A2:J2')
                merged_cell = worksheet['A2']
                merged_cell.fill = header_fill
                merged_cell.font = Font(color='FFFFFF', bold=True, name='맑은 고딕', size=20)
                merged_cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # 병합된 셀에 텍스트 입력 (동적 인원수 사용)
                merged_cell.value = f"{product_name} {platform} 체험단 ({num_people}명)"
                
                # A2-J2 행 높이 설정
                worksheet.row_dimensions[2].height = 50
                
                # 8행(실제 4행)에 스타일 적용
                for col in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=4, column=col)
                    cell.fill = header_fill
                    cell.font = white_bold_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # '비고' 열의 데이터 영역만 빈 값으로 설정
                    if cell.value == '비고':
                        for row in range(5, worksheet.max_row + 1):  # 헤더 아래의 데이터 영역
                            data_cell = worksheet.cell(row=row, column=col)
                            data_cell.value = ''
                
                # 금액 열 찾기 (헤더에서 '금액' 찾기)
                money_col = None
                for col in range(1, worksheet.max_column + 1):
                    if worksheet.cell(row=4, column=col).value == '금액':
                        money_col = col
                        break
                
                if money_col:
                    # 마지막 데이터가 있는 행 찾기
                    last_data_row = 4  # 헤더행부터 시작
                    for row in range(4, worksheet.max_row + 1):
                        row_has_data = False
                        for col in range(1, worksheet.max_column + 1):
                            if worksheet.cell(row=row, column=col).value:
                                row_has_data = True
                                last_data_row = row
                        if not row_has_data and row > 4:  # 헤더 행 이후에 데이터가 없는 행을 만나면
                            break

                # 데이터 영역에만 테두리와 중앙 정렬 적용 (4행부터 마지막 데이터 행까지)
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                # 테두리와 정렬 적용
                for row in range(4, last_data_row + 1):
                    for col in range(1, worksheet.max_column + 1):
                        cell = worksheet.cell(row=row, column=col)
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                # 나머지 스타일 적용 (금액 열 서식, 합계, 개당 가격 등)
                if money_col:
                    # 금액 열 서식 적용
                    for row in range(5, last_data_row + 1):
                        cell = worksheet.cell(row=row, column=money_col)
                        if cell.value:
                            try:
                                numeric_value = int(str(cell.value).replace(',', ''))
                                cell.value = numeric_value
                                cell.number_format = '#,##0'
                            except ValueError:
                                continue

                    # 합계 행 추가 및 스타일 적용
                    sum_row = last_data_row + 2

                    # 합계 레이블 추가 (금액 열 왼쪽)
                    label_cell = worksheet.cell(row=sum_row, column=money_col - 1)
                    label_cell.value = "합계:"
                    label_cell.font = Font(bold=True)
                    label_cell.alignment = Alignment(horizontal='right', vertical='center')
                    
                    # 합계 값을 함수식으로 추가
                    sum_cell = worksheet.cell(row=sum_row, column=money_col)
                    sum_cell.value = f'=SUM({get_column_letter(money_col)}5:{get_column_letter(money_col)}{last_data_row})'
                    sum_cell.number_format = '#,##0'
                    
                    # 개당 가격 레이블 추가 (합계 행 아래)
                    per_unit_label_cell = worksheet.cell(row=sum_row + 1, column=money_col - 1)
                    per_unit_label_cell.value = "개당 가격:"
                    per_unit_label_cell.font = Font(bold=True)
                    per_unit_label_cell.alignment = Alignment(horizontal='right', vertical='center')
                    
                    # 개당 가격 값을 함수식으로 추가 (동적 인원수 사용)
                    per_unit_cell = worksheet.cell(row=sum_row + 1, column=money_col)
                    per_unit_cell.value = f'={get_column_letter(money_col)}{sum_row}/{num_people}'
                    per_unit_cell.number_format = '#,##0'
                    per_unit_cell.font = Font(bold=True)
                    per_unit_cell.alignment = Alignment(horizontal='right', vertical='center')
                    
                    # 합계 행과 개당 가격 행에 배경색 추가 (테두리 없음)
                    light_gray_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
                    for row in [sum_row, sum_row + 1]:
                        for col in range(1, worksheet.max_column + 1):
                            cell = worksheet.cell(row=row, column=col)
                            cell.fill = light_gray_fill
                    
                    # 합계 행과 개당 가격 행의 높이 설정
                    worksheet.row_dimensions[sum_row].height = 20
                    worksheet.row_dimensions[sum_row + 1].height = 20
            
            # 눈금선 숨기기
            worksheet.sheet_view.showGridLines = False
            
            # 나머지 셀들의 기본 스타일 설정
            default_font = Font(name='맑은 고딕', size=11)
            
            # 헤더 행과 데이터 영역을 제외한 모든 셀의 스타일 초기화
            for row_idx in range(1, worksheet.max_row + 1):
                if row_idx != 4 and row_idx != 2 and row_idx not in [sum_row, sum_row + 1]:  # 헤더 행, 병합된 행, 합계/개당 가격 행 제외
                    for col_idx in range(1, worksheet.max_column + 1):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.font = default_font
            
            print(f'파일이 성공적으로 저장되었습니다: {output_file}')
            
        except Exception as e:
            print(f'오류가 발생했습니다: {str(e)}')

    def copy_invoice_template(self, product_name, save_path):
        """견적서 샘플을 복사하고 내용을 수정하는 함수"""
        # 현재 스크립트 경로 가져오기
        script_dir = os.path.dirname(os.path.abspath(__file__))
        
        # 견적서 샘플 파일 경로
        template_path = os.path.join(script_dir, "견적서 샘플.xlsx")
        
        # 오늘 날짜 형식 지정
        today = datetime.now().strftime("%y%m%d")
        today_formatted = datetime.now().strftime("%y년 %m월 %d일")
        
        # 새 파일명 생성
        new_filename = f"{product_name}_실비차감내역_{today}.xlsx"
        
        # 새 파일 경로
        new_file_path = os.path.join(save_path, new_filename)
        
        # 플랫폼 가져오기
        platform = self.platform_entry.get() if self.platform_var.get() == "직접입력" else self.platform_var.get()
        
        # 파일 복사
        if os.path.exists(template_path):
            shutil.copy2(template_path, new_file_path)
            print(f"견적서 샘플이 복사되었습니다: {new_file_path}")
            
            # 복사된 파일 열기
            wb = openpyxl.load_workbook(new_file_path)
            ws = wb.active
            
            # 현황판 파일 경로 (방금 생성된 파일)
            current_file = os.path.join(save_path, f"현황판_{product_name}_{platform}리뷰체험단_고야앤드_{today}.xlsx")
            
            # 인원수와 합계액, 개당 가격 계산
            num_people = 0
            total_amount = 0
            avg_price = 0
            
            if os.path.exists(current_file):
                # 현황판 파일에서 데이터 가져오기
                current_wb = openpyxl.load_workbook(current_file)
                current_ws = current_wb.active
                
                # 금액 열 찾기 (헤더에서 '금액' 찾기)
                money_col = None
                for col in range(1, current_ws.max_column + 1):
                    if current_ws.cell(row=4, column=col).value == '금액':
                        money_col = col
                        break
                
                if money_col:
                    # 마지막 데이터가 있는 행 찾기
                    last_data_row = 4  # 헤더행부터 시작
                    for row in range(4, current_ws.max_row + 1):
                        row_has_data = False
                        for col in range(1, current_ws.max_column + 1):
                            if current_ws.cell(row=row, column=col).value:
                                row_has_data = True
                                last_data_row = row
                        if not row_has_data and row > 4:  # 헤더 행 이후에 데이터가 없는 행을 만나면
                            break
                
                    # 데이터가 있는 행 수 계산 (5행부터 시작)
                    for row in range(5, last_data_row + 1):
                        # 제품 및 내역 열(C열)에 데이터가 있으면 카운트
                        if current_ws.cell(row=row, column=3).value:
                            num_people += 1
                        
                        # 금액 열에서 값 가져와서 합계 계산
                        cell_value = current_ws.cell(row=row, column=money_col).value
                        if cell_value:
                            try:
                                # 숫자나 문자열로 된 숫자 처리
                                if isinstance(cell_value, (int, float)):
                                    total_amount += cell_value
                                else:
                                    # 쉼표 제거 후 숫자로 변환
                                    numeric_value = int(str(cell_value).replace(',', ''))
                                    total_amount += numeric_value
                            except (ValueError, TypeError):
                                pass
                
                    # 개당 평균가 계산
                    if num_people > 0:
                        avg_price = total_amount // num_people
            
            # 추가 정보 입력
            # C3에 오늘 날짜 입력
            ws['C3'] = f'견적일 : {today_formatted}'
            
            # C4에 수신 입력
            ws['C4'] = '수 신 : 팔도'
            
            # C5에 담당 입력
            ws['C5'] = '담 당 : 신현빈'
            
            # C6에 상품 입력
            ws['C6'] = f'상 품 : {platform}리뷰체험단'
            
            # C11 셀 설정
            ws['C11'] = f'{platform}_{product_name}_상품가'
            
            # G11,H11 셀 병합 및 값 설정
            ws.merge_cells('G11:H11')
            ws['G11'] = 1
            
            # I11,J11 셀 병합 및 값 설정 (현황판에서 계산된 합계액)
            ws.merge_cells('I11:J11')
            ws['I11'] = total_amount
            
            # K11,L11 셀 병합 및 함수식 설정
            ws.merge_cells('K11:L11')
            ws['K11'] = '=G11*I11'
            
            # M11,N11 셀 병합 및 함수식 설정
            ws.merge_cells('M11:N11')
            ws['M11'] = '=K11*0.1'
            
            # O11 셀 설정 (현황판에서 계산된 개당 평균가)
            ws['O11'] = f'1개 평균가 {avg_price}원'
            
            # C12 셀 설정
            ws['C12'] = '구매리뷰어 체험단'
            
            # G12,H12 셀 병합 및 값 설정
            ws.merge_cells('G12:H12')
            ws['G12'] = num_people
            
            # I12,J12 셀 병합 및 값 설정
            ws.merge_cells('I12:J12')
            ws['I12'] = 8000
            
            # K12,L12 셀 병합 및 함수식 설정
            ws.merge_cells('K12:L12')
            ws['K12'] = '=G12*I12'
            
            # M12,N12 셀 병합 및 함수식 설정
            ws.merge_cells('M12:N12')
            ws['M12'] = '=K12*0.1'
            
            # O12 셀 설정
            ws['O12'] = ''
            
            # C13 셀 설정
            ws['C13'] = '배송비'
            
            # G13,H13 셀 병합 및 값 설정
            ws.merge_cells('G13:H13')
            ws['G13'] = num_people
            
            # I13,J13 셀 병합 및 값 설정
            ws.merge_cells('I13:J13')
            ws['I13'] = '-'
            
            # K13,L13 셀 병합 및 값 설정
            ws.merge_cells('K13:L13')
            ws['K13'] = '-'
            
            # M13,N13 셀 병합 및 값 설정
            ws.merge_cells('M13:N13')
            ws['M13'] = '-'
            
            # O13 셀 설정
            ws['O13'] = ''
            
            # 파일 저장
            wb.save(new_file_path)
            
        else:
            print(f"견적서 샘플 파일을 찾을 수 없습니다: {template_path}")
            self.status_label.config(text="견적서 샘플 파일을 찾을 수 없습니다.")

    def get_num_people(self):
        """현재 처리된 인원수 반환"""
        return getattr(self, 'num_people', 0)  # 속성이 없으면 0 반환

    def toggle_recipient_input(self, event=None):
        """담당자 콤보박스에서 '직접입력' 선택 시 입력창 표시/숨김"""
        if self.recipient_var.get() == "직접입력":
            self.recipient_entry.pack(side='left', padx=(5, 0), fill='x', expand=True)
            self.recipient_entry.focus()  # 입력창에 포커스 설정
        else:
            self.recipient_entry.pack_forget()
        
        # 버튼 상태 업데이트
        self.update_report_button_state()

    def update_report_button_state(self, event=None):
        """담당자 입력 여부에 따라 보고메일 버튼 상태 업데이트"""
        if self.recipient_var.get() == "직접입력":
            # 직접입력 모드일 때는 입력창의 내용 확인
            if self.recipient_entry.get().strip():
                self.report_email_button.config(state=tk.NORMAL)
            else:
                self.report_email_button.config(state=tk.DISABLED)
        else:
            # 콤보박스에서 선택했을 때는 항상 활성화
            self.report_email_button.config(state=tk.NORMAL)

    def get_recipient_name(self):
        """현재 선택된 담당자 이름 반환"""
        if self.recipient_var.get() == "직접입력":
            return self.recipient_entry.get().strip()
        else:
            return self.recipient_var.get()

    def show_report_email_from_button(self):
        """버튼에서 직접 호출될 때 사용하는 함수"""
        product = self.product_name.get()
        platform = self.platform_entry.get() if self.platform_var.get() == "직접입력" else self.platform_var.get()
        self.show_report_email(product, platform, self.get_num_people())

    def show_report_email(self, product=None, platform=None, num_people=0):
        # 새 창 생성
        email_window = tk.Toplevel(self.root)
        email_window.title("보고메일 양식")
        email_window.geometry("600x500")
        
        # 제목과 본문 프레임 생성
        subject_frame = tk.Frame(email_window)
        subject_frame.pack(fill='x', padx=10, pady=5)
        
        # 제목 표시
        tk.Label(subject_frame, text="제목: ").pack(side='left')
        subject_entry = tk.Entry(subject_frame, width=50)
        subject_entry.pack(side='left', fill='x', expand=True)
        
        # 제품명과 플랫폼이 제공되면 제목에 자동 입력
        if product and platform:
            subject_entry.insert(0, f"[고야앤드미디어] {product} {platform} 리뷰체험단 최종보고")
        else:
            subject_entry.insert(0, "[고야앤드미디어] 리뷰체험단 최종보고")
        
        # 본문 텍스트 영역
        body_text = tk.Text(email_window, height=20, width=60)
        body_text.pack(padx=10, pady=5)
        
        # 담당자 이름 가져오기
        recipient_name = self.get_recipient_name()
        
        # 기본 메일 내용 (동적으로 값 채우기)
        default_body = f"""안녕하세요 {recipient_name}님,
고야앤드미디어 신현빈입니다.

{product + ' ' + platform if product and platform else '제품명 플랫폼'} 리뷰체험단 결과보고 전달드립니다.
총 진행 인원은 {num_people}명으로,
상세내역은 첨부파일 내 현황판과 리뷰 캡쳐본 확인 부탁드립니다.

구매리뷰 체험단은 리뷰어별 상품구매 가격이 상이하여,
작업완료 후 실비정산 견적 함께 전달드리오니 참고부탁드립니다.
(해당 최종견적으로 패키지 차감 예정입니다.)


업무에 참고 부탁드립니다.

감사합니다.
신현빈 드림"""
        
        body_text.insert('1.0', default_body)
        
        # 복사 버튼
        button_frame = tk.Frame(email_window)
        button_frame.pack(pady=5)
        
        tk.Button(button_frame, 
                 text="제목 복사", 
                 command=lambda: self.copy_to_clipboard(subject_entry.get())
        ).pack(side='left', padx=5)
        
        tk.Button(button_frame, 
                 text="본문 복사", 
                 command=lambda: self.copy_to_clipboard(body_text.get('1.0', 'end-1c'))
        ).pack(side='left', padx=5)
        
        tk.Button(button_frame,
                 text="모두 복사",
                 command=lambda: self.copy_to_clipboard(f"제목: {subject_entry.get()}\n\n{body_text.get('1.0', 'end-1c')}")
        ).pack(side='left', padx=5)

    def copy_to_clipboard(self, text):
        self.root.clipboard_clear()
        self.root.clipboard_append(text)
        self.status_label.config(text="메일 내용이 클립보드에 복사되었습니다.")

def main():
    root = tk.Tk()
    app = StyleConverterGUI(root)
    root.mainloop()

if __name__ == '__main__':
    load_dotenv()
    main()
