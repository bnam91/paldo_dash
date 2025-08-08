'''
# STEP2.ë¸”ë¡œê±° ìµœê·¼ê¸€ í¬ë¡¤ë§ ver.1.2
- step 1ë‹¨ê³„ì—ì„œ ì¶”ì¶œí•œ ë¸”ë¡œê±°ë“¤ì˜ ìµœê·¼ ê¸€ì„ í¬ë¡¤ë§í•©ë‹ˆë‹¤.

### ë°©ë²•
- step 1ë‹¨ê³„ì—ì„œ ì¶”ì¶œí•œ ì—‘ì…€íŒŒì¼ì„ ê·¸ëŒ€ë¡œ ë„£ê³  ì‹¤í–‰í•˜ë©´ ë¨
- (a2)ëŠ” ì—°ê²°í•˜ì§€ ë§ê²ƒ. ì ë ¹í•  í‚¤ì›Œë“œëŠ” í•œë²ˆì— í•˜ë‚˜ì”©ì§„í–‰í•  ê²ƒ (ì…€ì´ ì•ˆë§ì•„ì„œ Cë ˆë²¨ì—ì„œ ì´ìƒí•˜ê²Œ ë‚˜ì˜´)
- ì²˜ìŒ ë¡œë”©ì‹œê°„ì´ ì œë²• ê±¸ë¦¬ëŠ” ì  ì°¸ê³ . 3ë¶„ ì´ë‚´ë¡œ ë˜ê¸´ í•¨

'''


import tkinter as tk
from tkinter import filedialog
import re
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import os
from datetime import datetime
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
import sys
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from auth import get_credentials

def select_excel_file():
    root = tk.Tk()
    root.withdraw()  # ë£¨íŠ¸ ì°½ ìˆ¨ê¸°ê¸°
    root.update()  # ìƒíƒœ ê°±ì‹ 

def select_excel_file():
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(title="ì—‘ì…€ íŒŒì¼ ì„ íƒ", filetypes=[("Excel files", "*.xlsx")])
    return file_path

def setup_webdriver():
    options = Options()
    options.headless = False
    driver = webdriver.Chrome(options=options)
    return driver


def extract_blog_ids(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    urls, names = [], []
    for row in range(1, ws.max_row + 1):
        url = ws['E' + str(row)].value
        name = ws['A' + str(row)].value
        if url and url.startswith("https://blog.naver.com/"):
            blog_id = url.split('/')[3]
            urls.append(f"https://blog.naver.com/PostList.naver?blogId={blog_id}&skinType=&skinId=&from=menu")
            names.append(name)
        time.sleep(1)
    return urls, names



def set_30_line_view(driver):
    select_box = driver.find_elements(By.CSS_SELECTOR, 'a.btn_select.pcol2._ListCountToggle._returnFalse')
    if select_box:
        select_box[0].click()
        time.sleep(1)
        thirty_lines_option = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, "//a[@data-value='30']"))
        )
        thirty_lines_option.click()
        time.sleep(2)
    else:
        open_list_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CLASS_NAME, 'btn_openlist.pcol2._toggleTopList._returnFalse'))
        )
        open_list_button.click()
        time.sleep(1)
        
        select_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, 'a.btn_select.pcol2._ListCountToggle._returnFalse'))
        )
        select_button.click()
        time.sleep(1)
        thirty_lines_option = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, "//a[@data-value='30']"))
        )
        thirty_lines_option.click()
        time.sleep(2)


def get_search_keyword():
    keyword = input("ê²€ìƒ‰í•  í‚¤ì›Œë“œë¥¼ ì…ë ¥í•˜ì„¸ìš”: ").strip()
    return keyword

def convert_to_date(date_str):
    try:
        # 'ë¶„ ì „', 'ì‹œê°„ ì „' ì²´í¬
        if 'ë¶„ ì „' in date_str or 'ì‹œê°„ ì „' in date_str:
            return datetime.now().strftime('%Y. %m. %d.')
            
        # '2025. 5. 19.' í˜•ì‹ ì²´í¬
        if re.match(r'\d{4}\. \d{1,2}\. \d{1,2}\.', date_str):
            return date_str
            
        # ê·¸ ì™¸ì˜ ê²½ìš° ì˜¤ëŠ˜ ë‚ ì§œ ë°˜í™˜
        return datetime.now().strftime('%Y. %m. %d.')
    except:
        return datetime.now().strftime('%Y. %m. %d.')

def get_sheet_names(service, spreadsheet_id):
    """êµ¬ê¸€ ì‹œíŠ¸ì˜ ëª¨ë“  ì‹œíŠ¸ ì´ë¦„ì„ ê°€ì ¸ì˜µë‹ˆë‹¤."""
    try:
        spreadsheet = service.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
        sheets = spreadsheet.get('sheets', [])
        sheet_names = [sheet['properties']['title'] for sheet in sheets]
        return sheet_names
    except Exception as e:
        print(f"ì‹œíŠ¸ ì´ë¦„ ê°€ì ¸ì˜¤ê¸° ì˜¤ë¥˜: {str(e)}")
        return []

def select_sheet(sheet_names):
    """ì‚¬ìš©ìê°€ ì‹œíŠ¸ë¥¼ ì„ íƒí•˜ë„ë¡ í•©ë‹ˆë‹¤."""
    print("\n=== ì‚¬ìš© ê°€ëŠ¥í•œ ì‹œíŠ¸ ëª©ë¡ ===")
    for i, name in enumerate(sheet_names, 1):
        print(f"{i}. {name}")
    
    while True:
        try:
            choice = int(input(f"\nì‹œíŠ¸ ë²ˆí˜¸ë¥¼ ì„ íƒí•˜ì„¸ìš” (1-{len(sheet_names)}): "))
            if 1 <= choice <= len(sheet_names):
                return sheet_names[choice - 1]
            else:
                print(f"1ë¶€í„° {len(sheet_names)} ì‚¬ì´ì˜ ìˆ«ìë¥¼ ì…ë ¥í•´ì£¼ì„¸ìš”.")
        except ValueError:
            print("ì˜¬ë°”ë¥¸ ìˆ«ìë¥¼ ì…ë ¥í•´ì£¼ì„¸ìš”.")

def show_preparation_guide():
    """ì¤€ë¹„ì‚¬í•­ì„ ì•ˆë‚´í•©ë‹ˆë‹¤."""
    print("\n" + "="*60)
    print("ğŸ“‹ ë¸”ë¡œê·¸ ì—…ë¡œë“œ íŠ¸ë˜í‚¹ ì¤€ë¹„ì‚¬í•­")
    print("="*60)
    print("\nğŸ” êµ¬ê¸€ ì‹œíŠ¸ ì¤€ë¹„ì‚¬í•­:")
    print("â€¢ Bì—´: ë¸”ë¡œê·¸ URL (ë„¤ì´ë²„ ë¸”ë¡œê·¸ ë§í¬)")
    print("â€¢ Cì—´: ì—…ë¡œë“œ ë§í¬ (ìë™ìœ¼ë¡œ ì±„ì›Œì§)")
    print("â€¢ Fì—´: ë¸”ë¡œê±° ì´ë¦„ (ì„ íƒì‚¬í•­)")
    print("â€¢ Iì—´: ì‘ì„±ì¼ (ìë™ìœ¼ë¡œ ì±„ì›Œì§)")
    print("\nğŸ“ ì‹œíŠ¸ êµ¬ì¡° ì˜ˆì‹œ:")
    print("Aì—´ | Bì—´(URL) | Cì—´(ì—…ë¡œë“œë§í¬) | Dì—´ | Eì—´ | Fì—´(ì´ë¦„) | Gì—´ | Hì—´ | Iì—´(ì‘ì„±ì¼)")
    print("-----|----------|----------------|-----|-----|----------|-----|-----|----------")
    print("     | https:// | (ìë™ì…ë ¥)     |     |     | ë¸”ë¡œê±°ëª…  |     |     | (ìë™ì…ë ¥)")
    print("\nâš ï¸  ì£¼ì˜ì‚¬í•­:")
    print("â€¢ Bì—´ì— ë„¤ì´ë²„ ë¸”ë¡œê·¸ URLì´ ìˆì–´ì•¼ í•©ë‹ˆë‹¤")
    print("â€¢ Cì—´ì´ ë¹„ì–´ìˆëŠ” í–‰ë§Œ ì²˜ë¦¬ë©ë‹ˆë‹¤")
    print("â€¢ í‚¤ì›Œë“œê°€ í¬í•¨ëœ ì²« ë²ˆì§¸ ê¸€ë§Œ ì—…ë°ì´íŠ¸ë©ë‹ˆë‹¤")
    print("â€¢ ê° ë¸”ë¡œê·¸ë‹¹ ìµœëŒ€ 30ê°œ ê¸€ì„ ê²€ìƒ‰í•©ë‹ˆë‹¤")
    print("\n" + "="*60)

def get_user_confirmation():
    """ì‚¬ìš©ì í™•ì¸ì„ ë°›ìŠµë‹ˆë‹¤."""
    while True:
        confirm = input("\nì¤€ë¹„ì‚¬í•­ì„ í™•ì¸í•˜ì…¨ë‚˜ìš”? (Y/N): ").strip().upper()
        if confirm == 'Y':
            return True
        elif confirm == 'N':
            return False
        else:
            print("Y ë˜ëŠ” Nì„ ì…ë ¥í•´ì£¼ì„¸ìš”.")

def update_sheet_with_link_and_date(service, spreadsheet_id, row_index, link, post_date, sheet_name):
    try:
        # ë‚ ì§œ í˜•ì‹ ë³€í™˜
        formatted_date = convert_to_date(post_date)
        
        # Cì—´ ì—…ë°ì´íŠ¸
        range_name_c = f"'{sheet_name}'!C{row_index}"
        body_c = {
            'values': [[link]]
        }
        service.spreadsheets().values().update(
            spreadsheetId=spreadsheet_id,
            range=range_name_c,
            valueInputOption='RAW',
            body=body_c
        ).execute()
        
        # Iì—´ ì—…ë°ì´íŠ¸ (ë³€í™˜ëœ ë‚ ì§œë¡œ)
        range_name_i = f"'{sheet_name}'!I{row_index}"
        body_i = {
            'values': [[formatted_date]]
        }
        service.spreadsheets().values().update(
            spreadsheetId=spreadsheet_id,
            range=range_name_i,
            valueInputOption='RAW',
            body=body_i
        ).execute()
        
    except Exception as e:
        print(f"ì‹œíŠ¸ ì—…ë°ì´íŠ¸ ì¤‘ ì˜¤ë¥˜ ë°œìƒ: {str(e)}")

def scrape_blog_data(driver, url, keyword, name, service, spreadsheet_id, row_index, sheet_name):
    driver.get(url)
    time.sleep(5)
    set_30_line_view(driver)
    try:
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'table.blog2_list'))
        )
        posts = driver.find_elements(By.CSS_SELECTOR, 'table.blog2_list tbody tr')[:30]
        data = []
        found_in_blog = False
        
        for post in posts:
            title_elements = post.find_elements(By.CSS_SELECTOR, 'span.ell2.pcol2')
            date_elements = post.find_elements(By.CSS_SELECTOR, 'div.wrap_td span.date.pcol2')
            link_elements = post.find_elements(By.CSS_SELECTOR, 'td.title a')
            if title_elements and date_elements and link_elements:
                title = title_elements[0].text
                date = date_elements[0].text
                link = link_elements[0].get_attribute('href')
                
                if keyword in title:
                    if not found_in_blog:
                        print(f"\ní‚¤ì›Œë“œ '{keyword}' ë°œê²¬")
                        found_in_blog = True
                        # ì‘ì„±ì¼ì„ íŒŒë¼ë¯¸í„°ë¡œ ì „ë‹¬
                        update_sheet_with_link_and_date(service, spreadsheet_id, row_index, link, date, sheet_name)
                    print(f"ì œëª©: {title}")
                    print(f"ì‘ì„±ì¼: {date}")
                    print(f"ë§í¬: {link}")
                    print("-" * 80)
                
                data.append((title, date, link))
                
        if not found_in_blog:
            print(f"{name}ì˜ ë¸”ë¡œê·¸ì—ì„œ '{keyword}' ê´€ë ¨ ê¸€ì´ ì—†ìŠµë‹ˆë‹¤.")
            
    except Exception as e:
        print(f"Error processing {url}: {str(e)}")
        data = []
    return data



def apply_excel_styles(ws):
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="FCD5B4", end_color="FCD5B4", fill_type="solid")
    center_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    grey_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 70
    ws.column_dimensions['D'].width = 16

    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = center_alignment
        cell.fill = header_fill
        cell.border = thin_border
    
    return grey_fill

def save_data_to_excel(urls, names, data_list, ws, grey_fill):
    for index, (url, name, data) in enumerate(zip(urls, names, data_list)):
        for title, date, link in data:
            row = [name, "", title, date, link]
            ws.append(row)
            if index % 2 == 1:
                for cell in ws[ws.max_row]:
                    cell.fill = grey_fill

def get_user_input():
    urls = []
    names = []
    print("ë¸”ë¡œê·¸ URLì„ ì…ë ¥í•˜ì„¸ìš” (ì…ë ¥ ì™„ë£Œì‹œ 'done' ì…ë ¥):")
    while True:
        url = input("URL: ").strip()
        if url.lower() == 'done':
            break
        if url.startswith("https://blog.naver.com/"):
            blog_id = url.split('/')[3]
            name = input("ë¸”ë¡œê±° ì´ë¦„: ").strip()
            urls.append(f"https://blog.naver.com/PostList.naver?blogId={blog_id}&skinType=&skinId=&from=menu")
            names.append(name)
        else:
            print("ì˜¬ë°”ë¥¸ ë„¤ì´ë²„ ë¸”ë¡œê·¸ URLì„ ì…ë ¥í•´ì£¼ì„¸ìš” (https://blog.naver.com/ë¡œ ì‹œì‘í•´ì•¼ í•©ë‹ˆë‹¤)")
    return urls, names

def print_blog_data(urls, names, data_list, keyword):
    print(f"\n=== '{keyword}' í‚¤ì›Œë“œ ê²€ìƒ‰ ê²°ê³¼ ===\n")
    found_any = False
    
    for url, name, data in zip(urls, names, data_list):
        found_in_blog = False
        for title, date, link in data:
            if keyword in title:
                if not found_in_blog:
                    print(f"\n[{name}ì˜ ë¸”ë¡œê·¸ì—ì„œ ë°œê²¬]")
                    found_in_blog = True
                    found_any = True
                print(f"ì œëª©: {title}")
                print(f"ì‘ì„±ì¼: {date}")
                print(f"ë§í¬: {link}")
                print("-" * 80)
    
    if not found_any:
        print(f"'{keyword}' í‚¤ì›Œë“œê°€ í¬í•¨ëœ ê¸€ì´ ì—†ìŠµë‹ˆë‹¤.")

def extract_sheet_id(input_str):
    """êµ¬ê¸€ ì‹œíŠ¸ URL ë˜ëŠ” IDì—ì„œ ì‹œíŠ¸ IDë¥¼ ì¶”ì¶œí•©ë‹ˆë‹¤."""
    # URL íŒ¨í„´ ë§¤ì¹­
    url_pattern = r'https://docs\.google\.com/spreadsheets/d/([a-zA-Z0-9-_]+)'
    match = re.search(url_pattern, input_str)
    
    if match:
        # URLì—ì„œ ID ì¶”ì¶œ
        return match.group(1)
    else:
        # URLì´ ì•„ë‹Œ ê²½ìš° ì…ë ¥ê°’ì´ IDë¼ê³  ê°€ì •
        # ID í˜•ì‹ ê²€ì¦ (ì•ŒíŒŒë²³, ìˆ«ì, í•˜ì´í”ˆ, ì–¸ë”ìŠ¤ì½”ì–´ë¡œë§Œ êµ¬ì„±)
        if re.match(r'^[a-zA-Z0-9-_]+$', input_str):
            return input_str
        else:
            raise ValueError("ìœ íš¨í•˜ì§€ ì•Šì€ êµ¬ê¸€ ì‹œíŠ¸ URL ë˜ëŠ” IDì…ë‹ˆë‹¤.")

def get_blog_data_from_sheet(service, spreadsheet_id_or_url, sheet_name):
    """êµ¬ê¸€ ì‹œíŠ¸ì—ì„œ ë¸”ë¡œê·¸ ë°ì´í„°ë¥¼ ê°€ì ¸ì˜µë‹ˆë‹¤."""
    try:
        # ì‹œíŠ¸ ID ì¶”ì¶œ
        spreadsheet_id = extract_sheet_id(spreadsheet_id_or_url)
        RANGE_NAME = f"'{sheet_name}'!B:F"  # ì„ íƒëœ ì‹œíŠ¸ì˜ Bì—´ë¶€í„° Fì—´ê¹Œì§€ ê°€ì ¸ì˜¤ë„ë¡ ìˆ˜ì •
        
        # ì‹œíŠ¸ ë°ì´í„° ê°€ì ¸ì˜¤ê¸°
        sheet = service.spreadsheets()
        result = sheet.values().get(
            spreadsheetId=spreadsheet_id,
            range=RANGE_NAME
        ).execute()
        
        values = result.get('values', [])
        if not values:
            print('ë°ì´í„°ê°€ ì—†ìŠµë‹ˆë‹¤.')
            return [], [], []
            
        urls = []
        names = []
        row_indices = []  # ì‹¤ì œ í–‰ ë²ˆí˜¸ë¥¼ ì €ì¥í•  ë¦¬ìŠ¤íŠ¸
        
        # ë°ì´í„° ì²˜ë¦¬ (ì²« ë²ˆì§¸ í–‰ì€ í—¤ë”ì´ë¯€ë¡œ ê±´ë„ˆë›°ê¸°)
        for i, row in enumerate(values[1:], start=2):
            print(f"\ní–‰ {i} ì²˜ë¦¬ ì¤‘:")
            print(f"í–‰ ë°ì´í„°: {row}")
            
            if len(row) >= 1:  # URLë§Œ ìˆì–´ë„ ì²˜ë¦¬ (Bì—´ë§Œ ìˆìœ¼ë©´ ë¨)
                print(f"ì—´ ê°œìˆ˜ ì¡°ê±´ í†µê³¼ (í˜„ì¬ {len(row)}ê°œ ì—´)")
                has_url = bool(row[0])  # Bì—´ì— URLì´ ìˆëŠ”ì§€
                print(f"Bì—´ URL ì¡´ì¬: {has_url} (ê°’: {row[0]})")
                
                # Cì—´ì´ ì—†ê±°ë‚˜ ë¹„ì–´ìˆëŠ” ê²½ìš° ì²˜ë¦¬
                is_cell_empty = len(row) <= 1 or not row[1] or row[1].strip() == ""
                print(f"Cì—´ ë¹„ì–´ìˆìŒ: {is_cell_empty} (ê°’: {row[1] if len(row) > 1 else 'ì—†ìŒ'})")
                
                if has_url and is_cell_empty:
                    print("URL ì²˜ë¦¬ ì‹œì‘")
                    blog_url = row[0]
                    # ëª¨ë°”ì¼ URLì„ PC URLë¡œ ë³€í™˜
                    if blog_url.startswith("https://m.blog.naver.com/"):
                        blog_url = blog_url.replace("https://m.blog.naver.com/", "https://blog.naver.com/")
                        # URLì—ì„œ ?tab=1 ê°™ì€ íŒŒë¼ë¯¸í„° ì œê±°
                        blog_url = blog_url.split('?')[0]
                    
                    # httpì™€ https ëª¨ë‘ ì²˜ë¦¬
                    if blog_url.startswith("http://blog.naver.com/") or blog_url.startswith("https://blog.naver.com/"):
                        # httpë¥¼ httpsë¡œ ë³€í™˜
                        if blog_url.startswith("http://blog.naver.com/"):
                            blog_url = blog_url.replace("http://blog.naver.com/", "https://blog.naver.com/")
                        
                        blog_id = blog_url.split('/')[3]
                        urls.append(f"https://blog.naver.com/PostList.naver?blogId={blog_id}&skinType=&skinId=&from=menu")
                        names.append(row[4] if len(row) > 4 else "")  # Fì—´(ì¸ë±ìŠ¤ 4)ì˜ ë¸”ë¡œê±° ì´ë¦„
                        row_indices.append(i)
                    # í”„ë¡œí† ì½œì´ ì—†ëŠ” blog.naver.com í˜•íƒœë„ ì²˜ë¦¬
                    elif blog_url.startswith("blog.naver.com/"):
                        blog_url = "https://" + blog_url
                        blog_id = blog_url.split('/')[3]  # https://blog.naver.com/eunshilys -> eunshilys
                        urls.append(f"https://blog.naver.com/PostList.naver?blogId={blog_id}&skinType=&skinId=&from=menu")
                        names.append(row[4] if len(row) > 4 else "")  # Fì—´(ì¸ë±ìŠ¤ 4)ì˜ ë¸”ë¡œê±° ì´ë¦„
                        row_indices.append(i)
                    else:
                        print("URL ì²˜ë¦¬ ê±´ë„ˆëœ€ (ë„¤ì´ë²„ ë¸”ë¡œê·¸ URLì´ ì•„ë‹˜)")
                else:
                    print("URL ì²˜ë¦¬ ê±´ë„ˆëœ€ (ì¡°ê±´ ë¶ˆì¶©ì¡±)")
            else:
                print(f"ì—´ ê°œìˆ˜ ë¶€ì¡± (í˜„ì¬ {len(row)}ê°œ ì—´)")
        
        print(f"\nìµœì¢… ì²˜ë¦¬ëœ URL ìˆ˜: {len(urls)}")
        return urls, names, row_indices
        
    except Exception as e:
        print(f"êµ¬ê¸€ ì‹œíŠ¸ ë°ì´í„° ì½ê¸° ì˜¤ë¥˜: {str(e)}")
        return [], [], []

def main():
    # ì¤€ë¹„ì‚¬í•­ ì•ˆë‚´
    show_preparation_guide()
    
    # ì‚¬ìš©ì í™•ì¸
    if not get_user_confirmation():
        print("í”„ë¡œê·¸ë¨ì„ ì¢…ë£Œí•©ë‹ˆë‹¤.")
        return
    
    # í‚¤ì›Œë“œ ì…ë ¥ ë°›ê¸°
    keyword = get_search_keyword()
    print(f"\n=== '{keyword}' í‚¤ì›Œë“œ ê²€ìƒ‰ ì‹œì‘ ===\n")
    
    # êµ¬ê¸€ ì‹œíŠ¸ ì„œë¹„ìŠ¤ ì´ˆê¸°í™”
    creds = get_credentials()
    service = build('sheets', 'v4', credentials=creds)
    
    # ì‹œíŠ¸ ID ì…ë ¥ ë°›ê¸°
    sheet_id_or_url = input("êµ¬ê¸€ ì‹œíŠ¸ URL ë˜ëŠ” IDë¥¼ ì…ë ¥í•˜ì„¸ìš”: ").strip()
    if not sheet_id_or_url:
        print("ì‹œíŠ¸ URL ë˜ëŠ” IDê°€ ì…ë ¥ë˜ì§€ ì•Šì•˜ìŠµë‹ˆë‹¤.")
        return
    
    try:
        # ì‹œíŠ¸ ID ì¶”ì¶œ
        spreadsheet_id = extract_sheet_id(sheet_id_or_url)
        
        # ì‚¬ìš© ê°€ëŠ¥í•œ ì‹œíŠ¸ ëª©ë¡ ê°€ì ¸ì˜¤ê¸°
        sheet_names = get_sheet_names(service, spreadsheet_id)
        if not sheet_names:
            print("ì‹œíŠ¸ë¥¼ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤.")
            return
        
        # ì‚¬ìš©ìê°€ ì‹œíŠ¸ ì„ íƒ
        selected_sheet = select_sheet(sheet_names)
        print(f"\nì„ íƒëœ ì‹œíŠ¸: {selected_sheet}")
        
        # êµ¬ê¸€ ì‹œíŠ¸ì—ì„œ ë¸”ë¡œê·¸ ë°ì´í„° ê°€ì ¸ì˜¤ê¸°
        urls, names, row_indices = get_blog_data_from_sheet(service, sheet_id_or_url, selected_sheet)
        if not urls:
            print("ì²˜ë¦¬í•  ë¸”ë¡œê·¸ ë°ì´í„°ê°€ ì—†ìŠµë‹ˆë‹¤.")
            return

        print(f"ì´ {len(urls)}ê°œì˜ ë¸”ë¡œê·¸ë¥¼ ê²€ìƒ‰í•©ë‹ˆë‹¤.")
        
        driver = setup_webdriver()
        data_list = []
        
        # ê° ë¸”ë¡œê·¸ë³„ë¡œ ì‹¤ì‹œê°„ ì²˜ë¦¬
        for url, name, row_index in zip(urls, names, row_indices):
            print(f"\n{name}ì˜ ë¸”ë¡œê·¸ ê²€ìƒ‰ ì¤‘...")
            data = scrape_blog_data(driver, url, keyword, name, service, spreadsheet_id, row_index, selected_sheet)
            data_list.append(data)
        
        driver.quit()
    except ValueError as e:
        print(f"ì˜¤ë¥˜: {str(e)}")
    except Exception as e:
        print(f"ì²˜ë¦¬ ì¤‘ ì˜¤ë¥˜ ë°œìƒ: {str(e)}")

if __name__ == "__main__":
    main()

    ## ì´ê²Œ ê±°ì˜ ìµœì¢…ì¼ë“¯
